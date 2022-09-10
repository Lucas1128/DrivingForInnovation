import cv2
import numpy as np
from pyzbar.pyzbar import decode
from openpyxl import load_workbook


loop = False

def run():
    print('Step 1')
    global loop
    loop = False
    wb = load_workbook(filename='Data.xlsx')
    sheet = wb['Sheet1']
    print('Step 2')

    def decoder(image):
        gray_img = cv2.cvtColor(image, 0)
        barcode = decode(gray_img)

        for obj in barcode:
            points = obj.polygon
            (x, y, w, h) = obj.rect
            pts = np.array(points, np.int32)
            pts = pts.reshape((-1, 1, 2))
            cv2.polylines(image, [pts], True, (0, 255, 0), 3)

            barcodeData = obj.data.decode("utf-8")
            barcodeType = obj.type
            string = "Data " + str(barcodeData) + " | Type " + str(barcodeType)

            cv2.putText(frame, string, (x, y), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 0, 0), 2)
            print("Barcode: " + barcodeData + " | Type: " + barcodeType)

            for row in range(1, sheet.max_row + 1):
                CellValue = sheet[row][0].value
                if int(barcodeData) == CellValue:
                    print('Match')
                    Data = []
                    Data.append(str(sheet[row][20].value)) #Alcohol by volume
                    #print('A')
                    Data.append(str(sheet[row][4].value)) #English Name
                    #print('B')
                    Data.append(str(sheet[row][7].value)) #French Name
                    #print('C')
                    Data.append(str(sheet[row][18].value)) #Brand Name
                    #print('D')

                    #print(Data)
                    global loop
                    loop = True

                    return Data



    cap = cv2.VideoCapture(0)
    print('Step 3')
    while loop != True:
        ret, frame = cap.read()
        values = decoder(frame)
        if values != None:
            cv2.destroyAllWindows()
            return values

        cv2.imshow('Image', frame)
        code = cv2.waitKey(10)

